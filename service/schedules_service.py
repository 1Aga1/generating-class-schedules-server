import datetime
import io
import os
import xlsxwriter

from openpyxl import load_workbook
from werkzeug import Response

from models import Schedules, ScheduleParams, GroupSubjects, Groups, Subjects, Teachers
from exceptions import ApiError

weekdays_en_ru = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье',
}

lessons_time = ['8:30 - 10:05', '10:15 - 11:50', '12:20 - 13:55', '14:05 - 15:40', '15:50 - 17:25', '17:35 - 19:10']
saturday_lessons_time = ['8:30 - 10:05', '10:15 - 11:50', '12:05 - 13:40', '13:50 - 15:25', '15:30 - 17:05', '17:15 - 18:50']


def create(date: str):
    schedule = Schedules(date=date)
    schedule.save()
    return schedule.get_dto()


def remove(schedule_id: int):
    schedule = Schedules.delete().where(Schedules.id == schedule_id)
    schedule.execute()


def edit(schedule_id: int, date: str):
    schedule = Schedules.get_or_none(Schedules.id == schedule_id)
    if not schedule:
        raise ApiError.BadRequest('Schedule not found')

    schedule.date = date
    schedule.save()

    return schedule.get_dto()


def get_schedules():
    return [schedule.get_dto() for schedule in Schedules.select().order_by(Schedules.date.asc())]


def get_schedule(schedule_id: int):
    schedule = Schedules.get_or_none(Schedules.id == schedule_id)
    if not schedule:
        raise ApiError.BadRequest('Schedule not found')

    return {
        **schedule.get_dto(),
        'params': [param.get_dto() for param in ScheduleParams.select().where(ScheduleParams.schedule == schedule).order_by(ScheduleParams.number.asc())]
    }


def upload_schedule(file):
    filename, file_extension = os.path.splitext(file.filename)
    if file_extension not in ['.xlsx', '.xls']:
        raise ApiError.BadRequest('Incorrect file format')

    work_book = load_workbook(file)

    for sheet_name in work_book.get_sheet_names():
        nigger = [int(item) for item in sheet_name.split('.')]
        schedule = create(datetime.date(nigger[2], nigger[1], nigger[0]).strftime('%Y-%m-%d'))

        sheet = work_book.get_sheet_by_name(sheet_name)

        for column in range(1, sheet.max_column + 1):

            if sheet.cell(row=1, column=column).value:
                group_name = sheet.cell(row=1, column=column).value

                group = Groups.get_or_none(course=int(group_name[0]), name=group_name)
                if not group:
                    group = Groups(course=int(group_name[0]), name=group_name)
                    group.save()

                for row in range(2, sheet.max_row+1):
                    cell_value = sheet.cell(row=row, column=column).value

                    if cell_value:
                        subjects = cell_value.split('/')

                        for data in subjects:
                            subject_w_teacher_w_office = data.split('-')
                            if subject_w_teacher_w_office[0] and subject_w_teacher_w_office[1] \
                                    and subject_w_teacher_w_office[2]:
                                teacher_fullname = subject_w_teacher_w_office[1]

                                teacher = Teachers.get_or_none(fullname=teacher_fullname)
                                if not teacher:
                                    teacher = Teachers(fullname=teacher_fullname)
                                    teacher.save()

                                subject = Subjects.get_or_none(name=subject_w_teacher_w_office[0],
                                                               teacher=teacher)
                                if not subject:
                                    subject = Subjects(name=subject_w_teacher_w_office[0],
                                                       teacher=teacher)
                                    subject.save()

                                group_subject = GroupSubjects.get_or_none(group=group, subject=subject)
                                if not group_subject:
                                    group_subject = GroupSubjects(group=group, subject=subject)
                                    group_subject.save()

                                schedule_param = ScheduleParams(schedule=schedule['id'], group=group, subject=subject,
                                                                number=row-1, office=subject_w_teacher_w_office[2])
                                schedule_param.save()

    return Response(status=200)


def change_visibility(schedule_id: int, visible: bool):
    schedule = Schedules.get_or_none(Schedules.id == schedule_id)
    if not schedule:
        raise ApiError.BadRequest('Schedule not found')

    schedule.visible = visible
    schedule.save()

    return schedule.get_dto()


def urtk_schedules():
    schedules = Schedules.select().where(Schedules.visible == True).order_by(Schedules.date.asc())

    urtk_schedules = {
        'dates': [schedule.date.strftime("%d.%m.%Y") for schedule in schedules],
        'lesson_number': 6,
        'courses': []
    }

    for course in range(1, 5):
        course_obj = {
            'name': course,
            'groups': []
        }

        groups = Groups.select().where(Groups.course == course)

        for group in groups:
            group_obj = {
                'name': group.name,
                'schedule': []
            }

            for schedule in schedules:
                schedule_params = ScheduleParams.select().where(
                    (ScheduleParams.schedule == schedule) & (ScheduleParams.group == group)) \
                    .order_by(ScheduleParams.number.asc())

                day = {
                    'date': schedule.date.strftime("%d.%m.%Y"),
                    'day': weekdays_en_ru[schedule.date.strftime("%A")],
                    'lessons': [{
                            'number': ' ',
                            'time': ' ',
                            'name': ' ',
                            'office': ' '
                        } for lesson in range(6)]
                }

                for index, param in enumerate(schedule_params):
                    subject = Subjects.get_or_none(Subjects.id == param.subject).get_dto()

                    lesson_name = f'{subject["name"]}'

                    if param.sub_group:
                        lesson_name += f' {param.sub_group}'

                    if param.first_half:
                        lesson_name = f'{param.first_half}/{lesson_name}'

                    lesson_name += f' {subject["teacher"]["fullname"]}'

                    if day['day'] != 'Суббота':
                        time = lessons_time[param.number - 1]
                    else:
                        time = saturday_lessons_time[param.number - 1]

                    if index > 0 and param.number == schedule_params[index - 1].number:
                        day['lessons'][param.number - 1]['name'] += f'/\n{lesson_name}'
                        day['lessons'][param.number - 1]['office'] += f'/{param.office}'
                    else:
                        day['lessons'][param.number - 1] = {
                            'number': param.number,
                            'time': time,
                            'name': f'{lesson_name}',
                            'office': f'{param.office}'
                        }

                group_obj['schedule'].append(day)

            course_obj['groups'].append(group_obj)

        urtk_schedules['courses'].append(course_obj)

    return urtk_schedules


def urtk_schedules_download():
    file_stream = io.BytesIO()

    document_name = 'Расписание.xlsx'
    workbook = xlsxwriter.Workbook(file_stream)

    urtk_schedule = urtk_schedules()

    with workbook:
        default_cell_format = workbook.add_format()
        default_cell_format.set_align('center')
        default_cell_format.set_align('vcenter')
        default_cell_format.set_text_wrap()
        default_cell_format.set_border()

        bold_format = workbook.add_format()
        bold_format.set_align('center')
        bold_format.set_align('vcenter')
        bold_format.set_text_wrap()
        bold_format.set_bold()

        bold_border_format = workbook.add_format()
        bold_border_format.set_align('center')
        bold_border_format.set_align('vcenter')
        bold_border_format.set_text_wrap()
        bold_border_format.set_bold()
        bold_border_format.set_border()

        date_format = workbook.add_format()
        date_format.set_align('center')
        date_format.set_align('vcenter')
        date_format.set_text_wrap()
        date_format.set_bold()
        date_format.set_border()
        date_format.set_rotation(90)

        for course in urtk_schedule['courses']:
            worksheet = workbook.add_worksheet(f'{course["name"]} курс')
            worksheet.merge_range(0, 0, 0, len(urtk_schedule['courses'])*3+1,
                                  f'Расписание занятий групп {course["name"]} курса на '
                                  f'{urtk_schedule["dates"][0]} - {urtk_schedule["dates"][len(urtk_schedule["dates"])-1]} '
                                  f'{urtk_schedule["dates"][0].split(".")[2]} года',
                                  bold_format
                                  )
            worksheet.write(1, 0, '', default_cell_format)

            row = 2
            for date in urtk_schedule["dates"]:
                worksheet.merge_range(row, 0, row+5, 0, date, date_format)
                row += 6

            group_col = 1
            for group in course['groups']:
                row = 1
                if row == 1:
                    worksheet.merge_range(row, group_col, row, group_col+2, group['name'], bold_border_format)


                row += 1
                for day in group['schedule']:
                    for lesson in day['lessons']:
                        worksheet.write(row, group_col, lesson['time'], default_cell_format)
                        worksheet.set_column(group_col, group_col, 6.5)

                        worksheet.write(row, group_col+1, lesson['name'], default_cell_format)
                        worksheet.set_column(group_col+1, group_col+1, 36)

                        worksheet.write(row, group_col+2, lesson['office'], default_cell_format)
                        worksheet.set_column(group_col+2, group_col+2, 7.6)

                        worksheet.set_row(row, 40)

                        row += 1

                group_col += 3

    file_stream.seek(0)

    return file_stream, document_name


def get_group_schedule(group_id: int):
    schedules = Schedules.select().where(Schedules.visible == True).order_by(Schedules.date.asc())

    group = Groups.get_or_none(Groups.id == group_id)

    if group:
        group_obj = {
            'name': group.name,
            'schedule': []
        }

        for schedule in schedules:
            schedule_params = ScheduleParams.select().where((ScheduleParams.schedule == schedule) & (ScheduleParams.group == group)) \
                .order_by(ScheduleParams.number.asc())

            day = {
                'date': schedule.date.strftime("%d.%m.%Y"),
                'day': weekdays_en_ru[schedule.date.strftime("%A")],
                'lessons': [{} for lesson in range(6)]
            }

            for index, param in enumerate(schedule_params):
                subject = Subjects.get_or_none(Subjects.id == param.subject).get_dto()

                lesson_name = f'{subject["name"]}'

                if param.sub_group:
                    lesson_name += f' {param.sub_group}'

                if param.first_half:
                    lesson_name = f'{param.first_half}/{lesson_name}'

                lesson_name += f' {subject["teacher"]["fullname"]}'

                if day['day'] != 'Суббота':
                    time = lessons_time[param.number-1]
                else:
                    time = saturday_lessons_time[param.number-1]

                if index > 0 and param.number == schedule_params[index-1].number:
                    day['lessons'][param.number-1]['name'] += f'/\n{lesson_name}'
                    day['lessons'][param.number-1]['office'] += f'/{param.office}'
                else:
                    day['lessons'][param.number-1] = {
                        'number': param.number,
                        'time': time,
                        'name': f'{lesson_name}',
                        'office': f'{param.office}'
                    }

            group_obj['schedule'].append(day)

        return group_obj

    raise ApiError.BadRequest('Group not found')